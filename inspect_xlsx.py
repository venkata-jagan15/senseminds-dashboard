import os
import sys

def main():
    print("Python version:", sys.version)
    
    # Try to install openpyxl and pandas locally if they aren't available
    try:
        import openpyxl
        print("openpyxl already installed")
    except ImportError:
        print("openpyxl not installed, attempting to install...")
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
        import openpyxl
        print("openpyxl installed successfully")
        
    try:
        import pandas as pd
        print("pandas already installed")
    except ImportError:
        print("pandas not installed, attempting to install...")
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "pandas"])
        import pandas as pd
        print("pandas installed successfully")

    xlsx_path = "c:\\Users\\Nartu Sindhusri\\OneDrive\\Desktop\\VOC\\Unit 1 Scrubbers & VOC six months data (1).xlsx"
    if not os.path.exists(xlsx_path):
        print(f"Error: file not found at {xlsx_path}")
        return

    # Inspect sheet names
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    print("Sheet names:", wb.sheetnames)
    
    # Read sheets
    for name in wb.sheetnames:
        df = pd.read_excel(xlsx_path, sheet_name=name)
        print(f"\n--- Sheet: {name} ---")
        print("Shape:", df.shape)
        print("Columns:", df.columns.tolist())
        print("Head:")
        print(df.head(5))
        print("Summary description:")
        print(df.describe(include='all'))

if __name__ == "__main__":
    main()
